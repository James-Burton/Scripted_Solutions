import arcpy,os,shutil,time,openpyxl,xlsxwriter,psutil
import pandas as pd
from xlsxwriter import Workbook
from openpyxl import load_workbook


## this function takes a list of values, or a value, and assigns it to the cell choosen. For lists, the cells below will be populated with each value in list.
## This function is used in report writer fucntion below.
def writeDetails(col,row,alist,sheet,xcldoc):
    ## loads data package
    book=load_workbook(xcldoc)
    ## sets the right sheet in data package
    sheet1=book[sheet]
    ## starts counter for row pasting iteration.
    ryy=0
    ## if input value is a lsit of values
    if type(alist) == list:
        ## for each item in list
        for yy in alist:
            ## set the current row
            rowiter=int(row+ryy)
            ## set the lists value to cell
            sheet1.cell(rowiter,col).value=yy
            ## save each paste. Can this be moved out of function?
            book.save(xcldoc)
            ## add one to row pasting iteration
            ryy+=1
    ## if value was not a lsit
    else:
        ## assign value to cell
        sheet1.cell(row,col).value=alist
        ## save.
        book.save(xcldoc)
    ## save a tiny bit of space by deleting ryy.
    del ryy

## this function writes all the input data sources to data package.
def reportBasics(username,ScaleData,TSA,HARVAUTH,OPENINGS,RDS,CUTBLOCK,RSRVS,aoiSubUnits,SkeenaD,col2,col3,col4,Today,report_out,ScriptName,ScriptContact,ScriptFor,ScriptV,ScriptP,GitLink, UniTM):
    ## load the xcel
    book=load_workbook(report_out)
    ## set the relevant sheets
    sheet1=book['MainPage']
    sheet4=book['RunSummary']
    ## set front page information
    sheet1['B3']=aoiSubUnits
    sheet1['B4']=col2
    ## set run summary basic information
    sheet4['B3']=username
    sheet4['B5']=Today
    ## record the path of input data
    sheet4['B10']=ScaleData
    sheet4['B11']=aoiSubUnits
    sheet4['B12']=TSA
    sheet4['B13']=RDS
    sheet4['B14']=OPENINGS
    sheet4['B15']=CUTBLOCK
    sheet4['B16']=HARVAUTH
    sheet4['B17']=SkeenaD
    sheet4['B18']=RSRVS
    ## identify what field is the primary aoi sub unit
    sheet4['B20']=col2
    ## secondary aoi sub unit
    sheet4['B21']=col3
    ## third aoi sub unit
    sheet4['B22']=col4
    ## details about the script
    sheet4['B31']=ScriptName
    sheet4['B32']=ScriptV
    sheet4['B33']=ScriptP
    sheet4['B34']=ScriptFor
    sheet4['B36']=ScriptContact
    sheet4['B35']=GitLink
    ## record UniLic count.
    sheet4['B38']=len(UniTM)
    book.save(report_out)


def MakeLicList(ScaleData,UniLic,tabularFF,UniTM,tabularTM,UniPE,tabularCuttingPE):

    ScaleData=r'NOT/REAL/PATH/ScaleData.xlsx'
    dataOG = pd.read_excel(ScaleData)
    #dataOG = dataOG.Licence.unique()
    UniLic=dataOG[tabularFF].tolist()
    UniTM=dataOG[tabularTM].tolist()
    UniPE=dataOG[tabularCuttingPE].tolist()
    print(len(UniLic))
    UniLic=reduce(lambda l, x: l.append(x) or l if x not in l else l, UniLic, [])
    y=0
    #UniLic=UniLic.encode('ascii')
    for x in UniLic:
        UniLic[y]=x.encode('ascii')
        y+=1
    y=0
    #UniLic=UniLic.encode('ascii')
    for x in UniTM:
        UniTM[y]=x.encode('ascii')
        y+=1
    y=0
    #UniLic=UniLic.encode('ascii')
    for x in UniPE:
        UniPE[y]=x.encode('ascii')
        y+=1
    print(len(UniLic))
    #print(UniLic)
    print(str((psutil.Process(os.getpid()).memory_info().rss / 1024 ** 2))+" MB in use.")

    del ScaleData


def append_to_excel(fpath, df, sheet_name):
    with pd.ExcelWriter(fpath, mode="A") as f:
        df.to_excel(f, sheet_name)

def DictDataFrame(dictDF,book,sheet5):
    dictDF=pd.DataFrame.from_dict( orient='index')
    dictDF.drop(0,axis=1)
    dictDF=pd.DataFrame.drop_duplicates(dictDF)
    wb2 = load_workbook('VnV_OutReport_rough24.xlsx')
    wb2.create_sheet('rex13')
    dictDF.to_excel('VnV_OutReport_rough24.xlsx','rex13')
    wb2.save('VnV_OutReport_rough24.xlsx')
    #dictDF.to_excel('VnV_OutReport_rough23.xlsx','TimberMarkDetails')
   # with pd.ExcelWriter('VnV_OutReport_rough12.xlsx') as writer:
        #dictDF.to_excel(writer,sheet_name='TimberMarkDetails')
    #append_to_excel('VnV_OutReport_rough24.xlsx', dictDF, 'TimberMarkDetails')
    del dictDF
    del MasterDictionary


def wrtmatchXL(inData,inDataCol,inList,TempContainer,outReportName):
    outReport=inData.loc[inData[inDataCol].isin(inList)]
    filepath = os.path.join(TempContainer,outReportName)
    outReport.to_excel(filepath, engine='xlsxwriter')
    del outReport
    print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')
    print(str(outReportName)+" written and now complete")
    print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

def MatchXL(inData,inDataCol,book,sheet,col_letter):
    outReport=inData.loc[inData[inDataCol].isin(inList)]


def wrtunmatchXL(inData,inDataCol,inList,TempContainer,outReportName):
    outReport=inData.loc[~inData[inDataCol].isin(inList)]
    filepath = os.path.join(TempContainer,outReportName)
    outReport.to_excel(filepath, engine='xlsxwriter')
    del outReport
    print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')
    print(str(outReportName)+" written and now complete")
    print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

