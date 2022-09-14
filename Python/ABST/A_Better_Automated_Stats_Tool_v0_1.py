
## TITLE: JAMES Tool or A Better AUtomated Status Tool [ABST]
##
## Authour: James "REX" Burton
##
## Date: Various Months in 2021
##
## PRE-REQS: 
## 1. Established permissions to access BCGW.sde supplied interanlly to BCGOV employees
## 2. ArcPy// ESRI Licencing// ArcMap 10.6 Server employing py2.7
## 3. pre-existent knowledge regarding business needs surrounding land use applications.
## 4. Must understand SQL, Feature Layer Management vs Feature Class Management, Da Cursor module, openpyxl knowledge.
##
## NOTES: 
## 1. This file has been stripped of values that may compromise internal security.
## 2. The spreadsheet report out functionality is not completed.
## 3. Bridge to auto-determine conlict resolution never supported or created.
## 4. No spelling checks.
##
##
## Import items needed 
import os, sys, arcpy, shutil, openpyxl, getpass, time
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def test_sde_connections():
    global bcgw
    bcgw = "Database Connections\\BCGW.sde"
    global bcgw_path
    bcgw_path = bcgw + "\\"
    # Test for sde connections as specified:
    print 'testing for BCGW connection....'
    if arcpy.Exists(bcgw) == True:
        print arcpy.AddMessage("    - BCGW connection is OK.")
    else:
        print "You must have an SDE connection named \"BCGW.sde\".  Please check your Database Connection and rerun this script."
        print "Operation terminated."
        sys.exit()

test_sde_connections()
## working space is where script is stored and where resultant datasets should be stored to reduce personal folder getting to hefty
working_space = r'[LOCAL_FILE_PATH_HERE]'

## AOI stand in
A1OTest = r'[LOCAL_FILE_PATH_HERE]'

## Common Variables
A1O = 'A1O' # Area of Interest

## set date/time
## set now as right this second now
rightnow= datetime.now()
## Today is a year,month,day string for temp gdb file name creation.
Today= rightnow.strftime('%Y_%m_%d')
## super temp string is for excel document creation so testing can occur in short succession while retaining ability to compare results between iterations\\not delete or overwrite xcel files
superTempString = rightnow.strftime("%m%d%Y%H%M%S")
## start program timmer
t0= time.clock()

# - # - # - # 
## OpenPYXL ins and outs and in-betweens
## In data
master_sheet = openpyxl.load_workbook(filename=r"[LOCAL_FILE_PATH_HERE].xlsx")['[specify worksheet here']

## Out Data 
report_book = openpyxl.Workbook()
report_out = r"[LOCAL_FILE_PATH_HERE]"+str(superTempString)+".xlsx"

## OPENPYXL formating
## highlight one works
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill("solid", fgColor="0099CCFF")

border_light1 = NamedStyle(name="border_light1")
bd11 = Side(style='thick', color="000000")
highlight.border = Border(left=bd11, top=bd11, right=bd11, bottom=bd11)

##highlight 1 does not work, was supposed to be for top row in multi conflict sections
highlight1 = NamedStyle(name="highlight1")
highlight.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill("solid", fgColor="0000CCFF")

## highlight2 does not work was supposed to be bottom row in multi conflict comparison
highlight2 = NamedStyle(name="highlight2")
highlight.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill("solid", fgColor="00CCFFFF")

def masterBufferList():
    ## set buffer values list to global for use outside of function
    global bufferValues
    ## make buffer list holder
    bufferValues = []
    ## add report title header higlight formatting to report book
    report_book.add_named_style(highlight)
    ## make new sheet in report
    report_sheet = 'Sheet1'
    ## title new sheet appropriately, turned off for debugging
    # report_sheeet.title = 'Main'
    ## specify buffer values column in master sheet
    buffer_value_column = 4
    ## for i in spreadsheet row range, starting after 1 so main header not picked up
    for i in range(2,master_sheet.max_row+1):
        if master_sheet.cell(i,1) is None:
            break
        else:                
            ## pick up values greater than 0 (means master sheet can have zero,0,or None values)
            if master_sheet.cell(i,buffer_value_column).value > 0:
                ## if value in list already, pass (do not add value to list)  
                if master_sheet.cell(i,buffer_value_column).value in bufferValues:
                    pass
                ## if value not in list, add it.
                else:
                    ## Weird ghost layer kept getting picked up on first input spreadsheet despite note exisitng. Tried to program it out, then made dataset2. Can leave here and make ghost layer file to skip if feelin keen.
                    if (master_sheet.cell(i,1).value) != '[LOCAL_FILE_PATH_HERE]':
                        try:
                            bufferValues.append(int(master_sheet.cell(i,buffer_value_column).value))
                        except:
                            print('bufferValue probably not an integer')
                            print('row : '+str(i))
                            print(str(master_sheet.cell(i,1).value))
masterBufferList()

## set AOI test shape up
arcpy.MakeFeatureLayer_management(A1OTest,A1O)

cat_names=['header','header end row','admin','admin end row','fn','first nation end','local','local end row']
cat_rows=[1,2,4,5,7,8,10,11]

# n=(cat_names.index('admin')+1)
# def updateRowCounts(n):
#     nn=n
#     for x in cat_rows[n:]:
#         cat_rows[nn]=x+1
#         nn+=1
#     report_sheet.insert_rows(cat_rows[n])


## Set title and what not on first sheet of out report
report_sheet = report_book.active
reportThis = 'Job Acceleration and Master Environmetnal Spreadsheet'
categ = 'header'
nn=(cat_names.index('header')+1)
n = nn
for xx in cat_rows[nn:]:
    cat_rows[n]=xx+1
    n+=1

nextrow=(cat_rows[cat_names.index(categ)+1])
report_sheet.insert_rows(nextrow)
report_sheet.cell(row=nextrow,column=1).value = reportThis
report_sheet.cell(row=nextrow,column=1).font = Font(bold=True)
report_sheet.cell(row=nextrow,column=1).style = highlight
report_book.save(filename=report_out)
reportThis = 'User:'
nn=(cat_names.index('header')+1)
n = nn
for xx in cat_rows[nn:]:
    cat_rows[n]=xx+1
    n+=1

nextrow=(cat_rows[cat_names.index(categ)+1])
report_sheet.insert_rows(nextrow)
report_sheet.cell(row=nextrow,column=1).value = reportThis
report_sheet.cell(row=nextrow,column=1).font = Font(bold=True)
reportThis = str(getpass.getuser())
report_sheet.cell(row=nextrow,column=2).value = reportThis
report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
report_book.save(filename=report_out)

reportThis = 'Date Run:'
nn=(cat_names.index('header')+1)
n = nn
for xx in cat_rows[nn:]:
    cat_rows[n]=xx+1
    n+=1

nextrow=(cat_rows[cat_names.index(categ)+1])
report_sheet.insert_rows(nextrow)
report_sheet.cell(row=nextrow,column=1).value = reportThis
report_sheet.cell(row=nextrow,column=1).font = Font(bold=True)

reportThis = Today
report_sheet.cell(row=nextrow,column=2).value = reportThis
report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
report_book.save(filename=report_out)



def SetUp():
    ## set variables
    global TempContainer
    global TempGDB
    global outWorkspace

    ## Sets the temporary container to be located in the raw data folder under the name Temp_[Today's date]
    TempContainer=os.path.join(working_space+ '\\Temp%s' % Today)

    ## Program checks if the Temp folder exists already
    if os.path.exists(TempContainer):

        ## If temp folder exists, remove the existing one and make a new one. This happens during script construction and should not normally
        ## be a problem if you run the tool once a day
        ## Script will fail at this point if temp gdb timestamps are still active. No solution to that.
        shutil.rmtree(TempContainer)
        os.makedirs(TempContainer)

    else:
        ## Normal function is to make the temp folder
        os.makedirs(TempContainer)

    ## Creates a temporary GDB in the temp folder.
    TempGDB = arcpy.CreateFileGDB_management(TempContainer, 'TempGDB.gdb', '10.0')

    ## Creates a vairable for quick programming to access the new temp gdb for arcpy env workspace
    outWorkspace = str(TempGDB)
    arcpy.env.workspace = outWorkspace

    ## this takes buffer values from column and buffers the AOI as needed
    for value in bufferValues:
        ## makes new feature class in temp gdb with AOI name and suffix buffer distance in meters
        tempBuff = A1O+str(value)
        ## makes the buffer distance dynamically
        tbv= str(value)+' Meters'
        ## execute our buffer and what not
        arcpy.Buffer_analysis(A1O,tempBuff,tbv,'FULL','ROUND','ALL')
        ## debugging statement for confirmation of completion
        print(str(tempBuff)+' was created with a '+tbv+' buffer :)')
        ## variables usually deleted after function but good habit to delete what not needed
        del tbv,tempBuff

SetUp()



###################################################################################################################################
## now the fun part
## determine query fields in spread sheet
## set to first five columns for basic tests

start_col = 0 # 'E' column index
end_col = 4# 'L' column index

## determine datasource filter, probably can make one filter and if equals bcgw then yes, else local source
sourceFilter1 = ['WHSE','spatialfiles']
#sourceFilter2 = 'spatialfiles'

## row by row, process conflicts
for ir in range(2, master_sheet.max_row+1):
    ## x represents the assumption there are YES multiple attributes to summarize
    x=1
    ## reset sql statement
    sqlStatement = ''
    ## represents the standard temp layer name, reset for each data source
    xQM = 'xQM'    
    ## queryme is an empty string to hold data source combination, reset for each data source
    queryme = ''
    ## try to delete temp layer (at top right now but should be move to bottom once function is complete.)
    try:
        arcpy.Delete_management(xQM)
    except:
        ## debugging statement
        print('first pass, xqm not created')

    ## get all values in row
    row = [cell.value for cell in master_sheet[ir][start_col:end_col+1]]
    categ = str(master_sheet.cell(ir,13).value)
    print categ
    ## for reference
        ## cell 0 is data source
        ## cell 1 is layer name/alis
        ## cell 3 is action, this is how many actions are applied to the datasource to avoid re-processing during analysis
            ## action = 1 means simple intersect
            ## action = 2 means intersect, buffer, and intersect again (example, does it cross a stream and then is it near any streams)
            ## action = 3 does not exisit but I reckon if there are two seperate sql statements for a layer then we can filter that properly with this field
        ## cell 4 is buffer value if applicable
        ## cell 5 is sql statement if applicable
        ## cell 6 to 12 is five fields to summarize and one field for map label

    ## debugging confirmation of data read
    print(row)

    ## set data source
    dataSource1 = str(row[0])

    ## quick check to make sure not at bottom of list
    if dataSource1 == 'None':
        sys.exit('Hit Bottom of List')

    ## set layer name
    layerName1 = str(row[1])
    
    ## get action count for datasource
    actionCount = row[2]
    
    ## if sql statement is present for this data source, assign it here
    if row[4] != None or str(row[4]) != 'None':
        sqlStatement = str(row[4])
    
    ## Determine attribute fields to summarize, make into list for cursor searching
    ## reset attributes summary for each datasource
    attributes_Summary = []
    ## basically, in row:in columns 6-12, get values
    for ic in range(6,master_sheet.max_column-1):
        ## as long as value is not None it will pick it up
        if master_sheet.cell(ir,ic).value != None: 
            ## if value already in the summary list, skip it 
            if master_sheet.cell(ir,ic).value in attributes_Summary:
                pass
        ## if value not in attribute sumamry, add it to list
            else:
                attributes_Summary.append(str(master_sheet.cell(ir,ic).value))
    print attributes_Summary
    ## if only one field found to summarize, remove article from list and make simple string, also set x counter to zero to represent list dissolved to string
    if len(attributes_Summary) == 1:
        attributes_Summary = attributes_Summary[0]
        x=0
    
    if len(attributes_Summary) ==0:
        x=6

    ## if data source evidently comes from bcgw
    if sourceFilter1[0].lower() in dataSource1.lower():
        ## query me becomes bcgw full path with data source
        queryme= str(bcgw_path+dataSource1)
    elif sourceFilter1[1].lower() in dataSource1.lower():
        queryme = dataSource1
    else:
        categ = 'header'
        nn=(cat_names.index('header')+1)
        n = nn
        for xx in cat_rows[nn:]:
            cat_rows[n]=xx+1
            n+=1

        nextrow=(cat_rows[cat_names.index(categ)+1])
        report_sheet.insert_rows(nextrow)
        report_sheet.cell(row=nextrow,column=1).value = layerName1+' does not exist!!!!!!!!!!!!'
        report_sheet.cell(row=nextrow,column=1).font = Font(bold=True)
        report_book.save(filename=report_out)
        pass

    next_space = report_sheet.max_row+2
    ## \/\/\/\/\/\/\/\/\/ make a temp layer name from queryme ##### IS THIS USED ANYWHERE?????????????????????????????????????????????????????????
    temp_layer = str('temp'+queryme)
    ## first make sure item exists
    if arcpy.Exists(queryme):
        print(queryme+' exists')
        ## make feature layer 
        arcpy.MakeFeatureLayer_management(queryme,xQM)
        field_names = [f.name for f in arcpy.ListFields(xQM)]
        print field_names
        print('was field names')
        ## make a list of field names in layer, this was to verify names in some layers during testing
        if x==6:
            field_names = [f.name for f in arcpy.ListFields(xQM)]
            xy=0
            while xy < 6:
                attributes_Summary.append(field_names[xy])
                xy+=1

        ## simple intersect between AOI and data source
        #pcr_test = str(arcpy.GetCount_management(xQM))
        #print(pcr_test + ' base records in '+xQM+' to select from')
        arcpy.SelectLayerByLocation_management(xQM,'intersect',A1O,'','NEW_SELECTION')
        #pcr_test = str(arcpy.GetCount_management(xQM))
        #print(pcr_test + ' records selected in '+xQM+' after location intersect')
        ## if there is a sql to apply, apply it here 
        if len(sqlStatement) > 1:
            arcpy.SelectLayerByAttribute_management(xQM, 'SUBSET_SELECTION',sqlStatement)
            #pcr_test = str(arcpy.GetCount_management(xQM))
            #print(pcr_test + ' records selected in '+xQM+' after location intersect and '+sqlStatement+' was used as sql select')
        
        iiii=0
        removed_dupe_conflicts = []


        with arcpy.da.SearchCursor(xQM, attributes_Summary) as cursor:
            for row33 in cursor:
                iii=0
                hl_pattern = 0
                if x == 0:
                    reportThis = attributes_Summary+' = ' +str(row33) 
                    if len(removed_dupe_conflicts) == 0:
                        report_sheet = report_book.active
                        nn=(cat_names.index(categ)+1)
                        n = nn
                        for xx in cat_rows[nn:]:
                            cat_rows[n]=xx+1
                            n+=1  
                        nn=(cat_names.index(categ)+1)
                        n = nn
                        for xx in cat_rows[nn:]:
                            cat_rows[n]=xx+1
                            n+=1
                        nextrow=(cat_rows[cat_names.index(categ)+1])
                        report_sheet.insert_rows(nextrow)
                        report_sheet.cell(row=nextrow,column=1).value= layerName1+' '+categ
                        report_sheet.cell(row=nextrow,column=2).value = reportThis
                        report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                        report_book.save(filename=report_out)
                        next_space += 1
                        removed_dupe_conflicts.append(reportThis)
                    else:
                        if reportThis in removed_dupe_conflicts:
                            pass
                        else:
                            nn=(cat_names.index(categ)+1)
                            n = nn
                            for xx in cat_rows[nn:]:
                                cat_rows[n]=xx+1
                                n+=1                           
                            nextrow=(cat_rows[cat_names.index(categ)+1])
                            report_sheet.insert_rows(nextrow)
                            report_sheet.cell(row=nextrow,column=2).value = reportThis
                            report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                            report_book.save(filename=report_out)

                            next_space += 1
                            removed_dupe_conflicts.append(reportThis)
                else:
                    while iii <= len(attributes_Summary)-1:
                        if iii ==0:
                            reportThis = str(attributes_Summary[iii])+' = '+str(row33[iii])
                            iii+=1
                        else:
                            reportThis = reportThis+'; '+str(attributes_Summary[iii])+' = '+str(row33[iii])
                            iii+=1
                    if len(removed_dupe_conflicts) == 0:
                        report_sheet = report_book.active
                        nn=(cat_names.index(categ)+1)
                        n = nn
                        for xx in cat_rows[nn:]:
                            cat_rows[n]=xx+1
                            n+=1  
                        nn=(cat_names.index(categ)+1)
                        n = nn
                        for xx in cat_rows[nn:]:
                            cat_rows[n]=xx+1
                            n+=1
                        nextrow=(cat_rows[cat_names.index(categ)+1])
                        report_sheet.insert_rows(nextrow)
                        report_sheet.cell(row=nextrow,column=1).value= layerName1+' '+categ
                        report_sheet.cell(row=nextrow,column=2).value = reportThis
                        report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                        report_book.save(filename=report_out)
                        next_space += 1
                        removed_dupe_conflicts.append(reportThis)
                    else:
                        if reportThis in removed_dupe_conflicts:
                            pass
                        else:
                            nn=(cat_names.index(categ)+1)
                            n = nn
                            for xx in cat_rows[nn:]:
                                cat_rows[n]=xx+1
                                n+=1                           
                            nextrow=(cat_rows[cat_names.index(categ)+1])
                            report_sheet.insert_rows(nextrow)
                            report_sheet.cell(row=nextrow,column=2).value = reportThis
                            report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                            report_book.save(filename=report_out)
                            next_space += 1
                            removed_dupe_conflicts.append(reportThis)
                

        if actionCount == 2:
            next_space = report_sheet.max_row+2
            row_4 = int(row[3])
            arcpy.env.workspace = outWorkspace
            buffA1O = A1O+str(row_4)
            layerName1 = layerName1+' within '+str(row_4)+' Meters.'
            print(layerName1)
            arcpy.SelectLayerByLocation_management(xQM,'intersect',buffA1O,'','NEW_SELECTION')
            if len(sqlStatement) > 1:
                arcpy.SelectLayerByAttribute_management(xQM, 'SUBSET_SELECTION',sqlStatement)
            iiii=0
            removed_dupe_conflicts=[]
            with arcpy.da.SearchCursor(xQM, attributes_Summary) as cursor:
                for row33 in cursor:
                    iii=0
                    hl_pattern = 0
                    if x == 0:
                        reportThis = attributes_Summary+' = ' +str(row33) 
                        if len(removed_dupe_conflicts) == 0:
                            report_sheet = report_book.active
                            nn=(cat_names.index(categ)+1)
                            n = nn
                            for xx in cat_rows[nn:]:
                                cat_rows[n]=xx+1
                                n+=1  
                            nn=(cat_names.index(categ)+1)
                            n = nn
                            for xx in cat_rows[nn:]:
                                cat_rows[n]=xx+1
                                n+=1
                            nextrow=(cat_rows[cat_names.index(categ)+1])
                            report_sheet.insert_rows(nextrow)

                            report_sheet.cell(row=nextrow,column=1).value= layerName1+' '+categ                     
                            report_sheet.cell(row=nextrow,column=2).value = reportThis
                            report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                            report_book.save(filename=report_out)
                            next_space += 1
                            removed_dupe_conflicts.append(reportThis)
                        else:
                            if reportThis in removed_dupe_conflicts:
                                pass
                            else:
                                nn=(cat_names.index(categ)+1)
                                n = nn
                                for xx in cat_rows[nn:]:
                                    cat_rows[n]=xx+1
                                    n+=1
                            
                                nextrow=(cat_rows[cat_names.index(categ)+1])
                                report_sheet.insert_rows(nextrow)
                                report_sheet.cell(row=nextrow,column=2).value = reportThis
                                report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                                report_book.save(filename=report_out)
                                next_space += 1
                                removed_dupe_conflicts.append(reportThis)
                    else:
                        while iii <= len(attributes_Summary)-1:
                            if iii ==0:
                                reportThis = str(attributes_Summary[iii])+' = '+str(row33[iii])
                                iii+=1
                            else:
                                reportThis = reportThis+'; '+str(attributes_Summary[iii])+' = '+str(row33[iii])
                                iii+=1
                        if len(removed_dupe_conflicts) == 0:
                            report_sheet = report_book.active
                            nn=(cat_names.index(categ)+1)
                            n = nn
                            for xx in cat_rows[nn:]:
                                cat_rows[n]=xx+1
                                n+=1                              
                            nn=(cat_names.index(categ)+1)
                            n = nn
                            for xx in cat_rows[nn:]:
                                cat_rows[n]=xx+1
                                n+=1  

                            nextrow=(cat_rows[cat_names.index(categ)+1])
                            report_sheet.insert_rows(nextrow)
                            report_sheet.cell(row=nextrow,column=1).value= layerName1+' '+categ
                            report_sheet.cell(row=nextrow,column=2).value = reportThis
                            report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                            report_book.save(filename=report_out)
                            next_space += 1
                            removed_dupe_conflicts.append(reportThis)
                        else:
                            if reportThis in removed_dupe_conflicts:
                                pass
                            else:
                                nn=(cat_names.index(categ)+1)
                                n = nn
                                for xx in cat_rows[nn:]:
                                    cat_rows[n]=xx+1
                                    n+=1
                            
                                nextrow=(cat_rows[cat_names.index(categ)+1])
                                report_sheet.insert_rows(nextrow)
                                report_sheet.cell(row=nextrow,column=2).value = reportThis
                                report_sheet.cell(row=nextrow,column=2).font = Font(bold=True)
                                report_book.save(filename=report_out)
                                removed_dupe_conflicts.append(reportThis)
                                next_space += 1
                    hl_pattern += 1 
    else:
        categ = 'header'
        nn=(cat_names.index('header')+1)
        n = nn
        for xx in cat_rows[nn:]:
            cat_rows[n]=xx+1
            n+=1

        nextrow=(cat_rows[cat_names.index(categ)+1])
        report_sheet.insert_rows(nextrow)
        report_sheet.cell(row=nextrow,column=1).value = layerName1+' does not exist! from error 2 !!!!!!!!!!!'
        report_sheet.cell(row=nextrow,column=1).font = Font(bold=True)
        report_book.save(filename=report_out)
        pass  

borderThick = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick')) #

borderTop = Border(top=Side(style='thick'))

borderBottom = Border(bottom=Side(style='thick'))

borderLeft = Border(left=Side(style='thick'))

borderRight = Border(right=Side(style='thick'))

borderTopLeft = Border(left=Side(style='thick'),
                    top=Side(style='thick'))

borderTopRight = Border(right=Side(style='thick'),
                    top=Side(style='thick'))

borderBottomLeft = Border(left=Side(style='thick'),
                    bottom=Side(style='thick')) #

borderBottomRight = Border(right=Side(style='thick'),
                    bottom=Side(style='thick')) #

borderSide = Side(border_style='thick')  # applies to sides of each cell
abc = 0
print len(cat_rows)
while abc < 8:

    # Set thin borders on entire sheet
    # a=cat_rows[abc]
    # abc +=1
    # b=cat_rows[abc]
    c=1
    d=3
    
    rowTop = int(cat_rows[abc])
    abc+=1
    rowBot = int(cat_rows[abc])
    rows = range(rowTop,rowBot)
    colLeft = int(c)
    colRight = int(d)
    columns = range(c,d)
    report_sheet = report_book.active
    sheet = report_sheet
    # Set thick outer borders
    for row in rows:
        for col in columns:
            sheet.cell(rowTop, col).border = borderTop
            sheet.cell(rowBot, col).border = borderBottom
            sheet.cell(row, colLeft).border = borderLeft
            sheet.cell(row, colRight).border = borderRight
            sheet.cell(rowTop,colLeft).border = borderTopLeft
            sheet.cell(rowTop,colRight).border = borderTopRight
            sheet.cell(rowBot,colLeft).border = borderBottomLeft
            sheet.cell(rowBot,colRight).border = borderBottomRight


    report_book.save(filename=report_out)
    abc+=1



t1 = time.clock() - t0
print('Time elapsed: ', (t1 - t0)) # CPU seconds elapsed (floating point)